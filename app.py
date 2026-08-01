import io
import json
import math
from dataclasses import dataclass
from datetime import date

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="WealthPath", page_icon="🏠", layout="wide")
st.markdown("""<style>
:root{
  --cream:#fffaf0; --paper:#ffffff; --green:#174b3a; --ink:#17231f;
  --muted:#5f6f68; --saffron:#d97706; --line:#e7dcc7; --soft-green:#edf7f2;
}

/* Fix the colour scheme independently of the viewer's Streamlit/browser theme. */
html, body, [data-testid="stAppViewContainer"], .stApp {
  background:var(--cream) !important;
  color:var(--ink) !important;
}
[data-testid="stMain"] .block-container {
  max-width:1320px;
  padding-top:3.2rem;
  padding-bottom:4rem;
}
[data-testid="stHeader"] {background:rgba(255,250,240,.96) !important;}
[data-testid="stToolbar"] {color:var(--ink) !important;}

/* Typography and ordinary Streamlit copy. */
h1, h2, h3, h4, h5, h6, p, label,
[data-testid="stMarkdownContainer"], [data-testid="stCaptionContainer"],
[data-testid="stWidgetLabel"], [data-testid="stMetricLabel"],
[data-testid="stMetricValue"], [data-testid="stMetricDelta"] {
  color:var(--ink) !important;
}
h1, h2, h3 {color:var(--green) !important;}
h1 {font-weight:800 !important; letter-spacing:-.035em;}
[data-testid="stCaptionContainer"] {color:var(--muted) !important;}

/* Tabs become the same warm, clear step navigation on every device. */
button[data-baseweb="tab"] {color:var(--muted) !important; font-weight:650;}
button[data-baseweb="tab"][aria-selected="true"] {color:var(--saffron) !important;}
[data-baseweb="tab-highlight"] {background-color:var(--saffron) !important;}
[data-baseweb="tab-border"] {background-color:var(--line) !important;}

/* Form controls: light surfaces with dark text, even under dark-mode defaults. */
[data-baseweb="input"] > div,
[data-baseweb="select"] > div,
[data-baseweb="base-input"],
[data-testid="stNumberInput"] input,
[data-testid="stDateInput"] input,
[data-testid="stTextInput"] input {
  background:var(--paper) !important;
  color:var(--ink) !important;
  border-color:var(--line) !important;
}
input, textarea {color:var(--ink) !important; -webkit-text-fill-color:var(--ink) !important;}
[data-testid="stNumberInputStepDown"], [data-testid="stNumberInputStepUp"] {
  background:#f6eedf !important; color:var(--green) !important;
}
[data-testid="stSlider"] {color:var(--ink) !important;}
[role="listbox"], [data-baseweb="popover"] {background:var(--paper) !important; color:var(--ink) !important;}
[role="option"] {color:var(--ink) !important;}

/* Tables and summaries. */
[data-testid="stDataFrame"], [data-testid="stDataEditor"] {
  border:1px solid var(--line); border-radius:14px; overflow:hidden; background:var(--paper);
}
div[data-testid="stMetric"] {
  background:var(--paper) !important; border:1px solid var(--line);
  border-radius:14px; padding:14px 16px; box-shadow:0 5px 18px rgba(47,37,19,.045);
}
[data-testid="stAlert"] {color:var(--ink) !important; border-radius:12px;}
.best {
  background:var(--soft-green); color:var(--ink) !important;
  border:1px solid #cfe6db; border-left:5px solid var(--green);
  padding:17px 18px; border-radius:12px; margin:.25rem 0 1rem;
}

/* Sidebar and primary actions. */
[data-testid="stSidebar"] {background:#f4ecdc !important; border-right:1px solid var(--line);}
[data-testid="stSidebar"] * {color:var(--ink);}
.stDownloadButton > button {
  background:var(--green) !important; color:white !important; border:0 !important;
  border-radius:10px !important; font-weight:700 !important;
}
.stDownloadButton > button:hover {background:#0f392b !important; color:white !important;}
hr {border-color:var(--line) !important;}

@media (max-width: 768px) {
  [data-testid="stMain"] .block-container {padding-top:2rem; padding-left:1rem; padding-right:1rem;}
  h1 {font-size:2rem !important;}
}
</style>""", unsafe_allow_html=True)

APP_DATA_VERSION = 1

DEFAULT_EXPENSES = pd.DataFrame({"Expense":["Household","Education","Insurance & medical","Travel","Other"],"Monthly amount":[35000,5000,5000,3000,2000]})
DEFAULT_ASSETS = pd.DataFrame({"Asset":["Bank / emergency fund","FD / debt","Equity / mutual funds","PF / NPS","Gold"],"Current value":[300000,100000,500000,400000,100000],"Use for own contribution %":[0,50,0,0,0],"Keep invested in wealth plan %":[0,0,100,0,0]})
DEFAULT_INVESTMENTS = pd.DataFrame({"Investment":["Equity mutual fund","Debt / FD"],"Current value":[400000,100000],"Monthly contribution":[5000,0],"Expected return %":[11.0,7.0],"Taxable":[True,True],"Tax rate %":[12.5,20.0],"Liquidity score (1-10)":[8,9]})

DATE_KEYS = {"loan_start"}
TABLE_KEYS = {"expenses_data","property_schedule_data","assets_data","bonus_sources_data","investments_data"}

def table_editor_key(table_key):
    """Use disposable widget keys because data_editor state is read-only."""
    revision=st.session_state.get("table_editor_revision",0)
    return f"{table_key}_editor_{revision}"

def stored_table(table_key,default):
    """Return the imported/latest table without sharing its key with a widget."""
    value=st.session_state.get(table_key,default)
    return value.copy() if isinstance(value,pd.DataFrame) else pd.DataFrame(value)

def remember_table(table_key,value):
    """Persist editor output under a non-widget session-state key."""
    st.session_state[table_key]=value.copy()
    return value

def json_safe(value):
    """Convert dates, numpy values and tables into portable JSON values."""
    if isinstance(value,pd.DataFrame):
        return [{str(k):json_safe(v) for k,v in row.items()} for row in value.to_dict("records")]
    if isinstance(value,(pd.Timestamp,date)):
        return value.isoformat()
    if isinstance(value,np.generic):
        return value.item()
    if pd.isna(value):
        return None
    return value

def planner_payload(values):
    return {"app":"WealthPath","data_version":APP_DATA_VERSION,"saved_at":pd.Timestamp.now().isoformat(),"inputs":{k:json_safe(v) for k,v in values.items()}}

def payload_to_csv(payload):
    rows=[]
    for field,value in payload.get("inputs",{}).items():
        section="Tables" if field in TABLE_KEYS else "Inputs"
        rows.append({"Section":section,"Field":field,"Value":json.dumps(value,ensure_ascii=False)})
    return pd.DataFrame(rows).to_csv(index=False).encode("utf-8-sig")

def csv_to_payload(raw):
    frame=pd.read_csv(io.BytesIO(raw))
    required={"Section","Field","Value"}
    if not required.issubset(frame.columns):
        raise ValueError("CSV must contain Section, Field and Value columns. Please use the provided template.")
    inputs={}
    for _,row in frame.iterrows():
        field=str(row["Field"]).strip()
        if not field or field.lower()=="nan":
            continue
        raw_value=row["Value"]
        if pd.isna(raw_value):
            continue
        try:
            inputs[field]=json.loads(str(raw_value))
        except json.JSONDecodeError as exc:
            raise ValueError(f"Invalid Value for '{field}'. Keep text in JSON quotes, for example \"Balanced plan\".") from exc
    return {"app":"WealthPath","data_version":APP_DATA_VERSION,"inputs":inputs}

def apply_import(payload):
    if not isinstance(payload,dict) or not isinstance(payload.get("inputs"),dict):
        raise ValueError("This is not a valid WealthPath backup file.")
    if payload.get("app") not in (None,"WealthPath"):
        raise ValueError("This file was not created for WealthPath.")
    inputs=payload["inputs"]
    imported_tables={}
    for key,value in inputs.items():
        if key in TABLE_KEYS:
            if not isinstance(value,list):
                raise ValueError(f"'{key}' must contain table rows.")
            table=pd.DataFrame(value)
            if "Expected date" in table.columns:
                table["Expected date"]=pd.to_datetime(table["Expected date"],errors="coerce").dt.date
            imported_tables[key]=table
        elif key in DATE_KEYS:
            parsed=pd.to_datetime(value,errors="coerce")
            if pd.isna(parsed):
                raise ValueError(f"'{key}' contains an invalid date.")
            st.session_state[key]=parsed.date()
        else:
            st.session_state[key]=value
    # data_editor widget state cannot be assigned through Session State. Store
    # imported tables separately and use a new generation of widget keys.
    st.session_state.update(imported_tables)
    st.session_state["table_editor_revision"]=st.session_state.get("table_editor_revision",0)+1

def template_values():
    start=date.today().replace(day=1)
    d0=pd.Timestamp(start)
    return {
        "objective":"Balanced plan","leveraged":False,"run_mc":False,
        "salary":120000.0,"spouse":0.0,"growth":6.0,"existing_emi":0.0,"buffer_pct":10,
        "property_value":7500000.0,"loan":5500000.0,"charges":500000.0,"rate":8.5,
        "tenure_y":20,"annual_prepay":0.0,"loan_start":start,"repayment_mode":"Pre-EMI until final disbursement",
        "expenses_data":DEFAULT_EXPENSES,"assets_data":DEFAULT_ASSETS,"investments_data":DEFAULT_INVESTMENTS,
        "property_schedule_data":pd.DataFrame({"Payment / milestone":["Booking / initial payment","Construction / second payment","Registration / possession"],"Expected date":[d0.date(),(d0+pd.DateOffset(months=6)).date(),(d0+pd.DateOffset(months=12)).date()],"Own contribution":[800000.0,600000.0,600000.0],"Loan disbursement":[1100000.0,2200000.0,2200000.0]}),
        "min_liquidity":300000.0,
        "bonus_sources_data":pd.DataFrame({"Source":["Annual bonus / incentive"],"Expected date":[(d0+pd.DateOffset(months=5)).date()],"Amount":[0.0]})
    }

st.title("🏠 WealthPath — Home Loan & Wealth Optimizer")
st.caption("Compare loan prepayment and investing using your actual income, expenses, assets and investment assumptions.")

with st.expander("💾 Import, export or restore planner data",expanded=False):
    st.caption("CSV is convenient for bulk editing. JSON is the recommended complete backup. Imported values stay in this browser session; download a backup before clearing or closing the app.")
    uploaded=st.file_uploader("Import WealthPath data",type=["json","csv"],key="planner_import_file",help="Upload a WealthPath JSON backup or a CSV that follows the provided template.")
    if st.button("Import and restore data",disabled=uploaded is None,use_container_width=True):
        try:
            raw=uploaded.getvalue()
            if uploaded.name.lower().endswith(".json"):
                incoming=json.loads(raw.decode("utf-8-sig"))
            else:
                incoming=csv_to_payload(raw)
            apply_import(incoming)
            st.session_state["import_message"]="Data restored successfully. All calculations have been refreshed from the imported inputs."
            st.rerun()
        except Exception as exc:
            st.error(f"Import failed: {exc}")
    if st.session_state.pop("import_message",None):
        st.success("Data restored successfully. All calculations have been refreshed from the imported inputs.")
    sample=planner_payload(template_values())
    a,b=st.columns(2)
    a.download_button("Download CSV import template",payload_to_csv(sample),"wealthpath_import_template.csv","text/csv",use_container_width=True)
    b.download_button("Download sample JSON",json.dumps(sample,ensure_ascii=False,indent=2).encode("utf-8"),"wealthpath_sample_backup.json","application/json",use_container_width=True)

def money(x): return f"₹{x:,.0f}"

def emi(principal, annual_rate, months):
    r=annual_rate/1200
    return principal/months if r==0 else principal*r*(1+r)**months/((1+r)**months-1)

def effective_return(row):
    """Return a row's post-tax annual rate without trusting editor values."""
    gross=pd.to_numeric(row.get("Expected return %"),errors="coerce")
    tax=pd.to_numeric(row.get("Tax rate %"),errors="coerce")
    gross=0.0 if pd.isna(gross) else float(gross)/100
    tax=0.0 if pd.isna(tax) else float(np.clip(tax,0,100))/100
    taxable=row.get("Taxable",False)
    taxable=False if pd.isna(taxable) else bool(taxable)
    return gross*(1-tax if taxable else 1)

def weighted_return(inv):
    """Calculate the weighted post-tax return and ignore incomplete editor rows."""
    if not isinstance(inv,pd.DataFrame) or inv.empty:
        return 0.0

    data=inv.copy()
    for column in ["Current value","Monthly contribution","Expected return %","Tax rate %"]:
        if column not in data.columns:
            data[column]=0.0
        data[column]=pd.to_numeric(data[column],errors="coerce")
    if "Taxable" not in data.columns:
        data["Taxable"]=False

    # A blank expected-return cell means the row is not ready for projection.
    data=data.dropna(subset=["Expected return %"])
    if data.empty:
        return 0.0
    data[["Current value","Monthly contribution","Tax rate %"]]=data[["Current value","Monthly contribution","Tax rate %"]].fillna(0)
    weights=(data["Current value"].clip(lower=0)+data["Monthly contribution"].clip(lower=0)*12)
    valid=weights>0
    if not valid.any():
        return 0.0
    returns=data.loc[valid].apply(effective_return,axis=1).astype(float)
    return float(np.average(returns,weights=weights.loc[valid]))

def month_number(start, value):
    """Convert a calendar date into a 1-based projection month."""
    start=pd.Timestamp(start).replace(day=1)
    value=pd.Timestamp(value).replace(day=1)
    return max(1,(value.year-start.year)*12+value.month-start.month+1)

def normalise_schedule(frame, start, amount_col):
    """Return a clean dated schedule, including when the editor is empty."""
    data=frame.copy() if isinstance(frame,pd.DataFrame) else pd.DataFrame()
    if "Date" not in data.columns and "Expected date" in data.columns:
        data=data.rename(columns={"Expected date":"Date"})
    if "Date" not in data.columns:
        data["Date"]=pd.Series(dtype="datetime64[ns]")
    if amount_col not in data.columns:
        data[amount_col]=pd.Series(dtype="float64")
    data["Date"]=pd.to_datetime(data["Date"],errors="coerce")
    data[amount_col]=pd.to_numeric(data[amount_col],errors="coerce").fillna(0).clip(lower=0)
    data=data.dropna(subset=["Date"])
    data["Month"]=data["Date"].apply(lambda x:month_number(start,x))
    return data

@dataclass
class Result:
    name:str; interest:float; close_month:int; corpus:float; net_worth:float; liquidity:float; risk:float; rows:list

def simulate(name, loan, rate, tenure_m, payment, start_corpus, inv_r, invest_share,
             prepay_frequency=None, annual_growth=0, leveraged=False,
             disbursements=None, contribution_reserve=None, repayment_mode="Pre-EMI until final disbursement"):
    disbursements=disbursements or {1:loan}
    contribution_reserve=contribution_reserve or {}
    balance=0.0; invested=start_corpus+(loan if leveraged else 0); cash=0.0; total_interest=0; rows=[]
    monthly_r=inv_r/12; loan_r=rate/1200; close_month=tenure_m
    final_disbursement=max(disbursements)
    regular_emi=emi(loan,rate,tenure_m)
    for m in range(1,361):
        loan_draw=disbursements.get(m,0.0)
        balance+=loan_draw
        opening_balance=balance
        opening_investment=invested
        yearly_factor=(1+annual_growth/100)**((m-1)//12)
        capacity=payment*yearly_factor
        own_contribution_saving=min(capacity,contribution_reserve.get(m,0.0))
        capacity=max(0,capacity-own_contribution_saving)
        interest=balance*loan_r
        if balance<=0:
            scheduled=0
        elif repayment_mode=="Pre-EMI until final disbursement" and m<final_disbursement:
            scheduled=interest
        else:
            scheduled=regular_emi
        principal=max(0,min(balance,scheduled-interest))
        balance-=principal; total_interest+=interest
        do_prepay=(prepay_frequency=="monthly" or
                   prepay_frequency=="quarterly" and m%3==0 or
                   prepay_frequency=="annual" and m%12==0)
        prepay=capacity*(1-invest_share) if do_prepay and balance>0 else 0
        prepay=min(balance,prepay); balance-=prepay
        invest=capacity*invest_share
        if name=="Only EMI": cash+=capacity
        if balance<=0 and m>=final_disbursement and close_month==tenure_m: close_month=m
        emi_redirect=scheduled if balance<=0 else 0
        if balance<=0: invest += emi_redirect
        investment_growth=invested*monthly_r
        invested=invested*(1+monthly_r)+invest
        corpus=invested+cash
        if own_contribution_saving>0: action=f"Reserve {money(own_contribution_saving)} for the next own-contribution payment"
        elif prepay>0 and invest>0: action="Pay EMI, prepay and invest"
        elif prepay>0: action="Pay EMI and make prepayment"
        elif invest>0: action="Pay EMI and invest"
        elif scheduled>0: action="Pay regular EMI; retain surplus"
        else: action="Loan closed; continue wealth building"
        rows.append({"Month":m,"Action":action,"Loan disbursed":loan_draw,
                     "Own contribution saving":own_contribution_saving,"Opening loan balance":opening_balance,
                     "Scheduled EMI":min(scheduled,opening_balance+interest),"Interest":interest,
                     "Principal repaid":principal,"Prepayment":prepay,"Loan outstanding":max(0,balance),
                     "Opening investment value":opening_investment,"Investment":invest,
                     "Expected investment growth":investment_growth,"Investment corpus":corpus,
                     "EMI redirected after closure":emi_redirect,"Strategy capacity after contribution":capacity})
    property_equity=max(0,loan-balance)
    risk=invest_share*8+(2 if leveraged else 0)
    return Result(name,total_interest,close_month,corpus,corpus+property_equity-balance,invest_share*10,risk,rows)

with st.sidebar:
    st.header("Your priority")
    objective=st.selectbox("What matters most?",["Balanced plan","Close loan fastest","Maximise net worth","Lowest risk","Highest liquidity"],key="objective",
        help="This changes how strategies are ranked; it does not change the underlying calculations.")
    leveraged=st.toggle("Can loan proceeds be invested?",False,key="leveraged",help="Normally No for a home loan. Yes models a high-risk leveraged scenario.")
    run_mc=st.toggle("Run 10,000 market simulations",False,key="run_mc")

t1,t2,t3,t4=st.tabs(["1 · Income & expenses","2 · Property funding & loan","3 · Assets & investments","4 · Plan & results"])
with t1:
    st.subheader("Monthly cash flow")
    st.info("The tool calculates available surplus; you do not enter a planned investment amount.")
    c1,c2,c3=st.columns(3)
    salary=c1.number_input("Monthly in-hand income",0.0,value=120000.0,step=5000.0,key="salary",help="Household take-home income available before expenses and EMIs.")
    spouse=c2.number_input("Spouse/other monthly income",0.0,value=0.0,step=5000.0,key="spouse")
    growth=c3.number_input("Annual income growth %",0.0,30.0,6.0,0.5,key="growth",help="Raises future surplus; treated as an assumption, not guaranteed.")
    expenses=remember_table("expenses_data",st.data_editor(stored_table("expenses_data",DEFAULT_EXPENSES),num_rows="dynamic",use_container_width=True,key=table_editor_key("expenses_data")))
    existing_emi=st.number_input("Other loan EMIs",0.0,value=0.0,step=1000.0,key="existing_emi")
    buffer_pct=st.slider("Safety buffer (% of income)",0,30,10,key="buffer_pct",help="Protected cash not allocated to investment or prepayment.")
    total_income=salary+spouse; total_exp=float(expenses["Monthly amount"].sum())
    safe_surplus=max(0,total_income-total_exp-existing_emi-total_income*buffer_pct/100)
    st.metric("Calculated safe monthly surplus",money(safe_surplus),help="Income − expenses − other EMIs − safety buffer.")

with t2:
    st.subheader("Home loan")
    c1,c2,c3=st.columns(3)
    property_value=c1.number_input("Property value",0.0,value=7500000.0,step=100000.0,key="property_value")
    loan=c2.number_input("Loan amount",0.0,value=5500000.0,step=100000.0,key="loan")
    charges=c3.number_input("Registration and other charges",0.0,value=500000.0,step=50000.0,key="charges")
    rate=c1.number_input("Current interest rate %",0.0,30.0,8.5,0.05,key="rate",help="Monthly reducing-balance rate. Update this when a floating rate changes.")
    tenure_y=int(c2.number_input("Remaining tenure (years)",1,40,20,key="tenure_y"))
    annual_prepay=c3.number_input("Optional annual lump sum",0.0,value=0.0,step=10000.0,key="annual_prepay",help="Bonus or other amount available once each year, separate from monthly surplus.")
    loan_start=st.date_input("Loan/report start date",value=date.today().replace(day=1),key="loan_start",help="Used to convert Month 1, Month 2 and later actions into calendar months in the Excel report.")
    repayment_mode=st.radio("Repayment during staged disbursement",["Pre-EMI until final disbursement","Full EMI from first disbursement"],horizontal=True,key="repayment_mode",
        help="Pre-EMI pays only interest on the amount already disbursed. Full EMI starts principal repayment from the first drawdown. Confirm your bank's actual method.")
    base_emi=emi(loan,rate,tenure_y*12)
    own_contribution=max(0,property_value+charges-loan)
    a,b=st.columns(2); a.metric("Calculated EMI",money(base_emi)); b.metric("Own contribution",money(own_contribution))
    st.subheader("When will the property money be paid?")
    st.caption("Enter every expected tranche. The loan and your own money need not be paid on the first day.")
    d0=pd.Timestamp(loan_start)
    property_schedule_default=pd.DataFrame({
        "Payment / milestone":["Booking / initial payment","Construction / second payment","Registration / possession"],
        "Expected date":[d0.date(),(d0+pd.DateOffset(months=6)).date(),(d0+pd.DateOffset(months=12)).date()],
        "Own contribution":[own_contribution*.40,own_contribution*.30,own_contribution*.30],
        "Loan disbursement":[loan*.20,loan*.40,loan*.40]})
    property_schedule=remember_table("property_schedule_data",st.data_editor(stored_table("property_schedule_data",property_schedule_default),num_rows="dynamic",use_container_width=True,key=table_editor_key("property_schedule_data"),
        column_config={
            "Expected date":st.column_config.DateColumn(help="Estimated date on which this tranche must reach the builder/seller."),
            "Own contribution":st.column_config.NumberColumn(min_value=0,format="₹ %.0f",help="Money paid from savings—not borrowed money."),
            "Loan disbursement":st.column_config.NumberColumn(min_value=0,format="₹ %.0f",help="Loan amount expected from the bank on this date. Interest begins only after disbursement.")}))
    entered_own=float(pd.to_numeric(property_schedule["Own contribution"],errors="coerce").fillna(0).sum())
    entered_loan=float(pd.to_numeric(property_schedule["Loan disbursement"],errors="coerce").fillna(0).sum())
    q1,q2=st.columns(2)
    q1.metric("Own contribution scheduled",money(entered_own),delta=money(entered_own-own_contribution))
    q2.metric("Loan disbursement scheduled",money(entered_loan),delta=money(entered_loan-loan))
    if abs(entered_own-own_contribution)>1 or abs(entered_loan-loan)>1:
        st.warning("The dated tranches must total the calculated own contribution and the sanctioned loan amount before the recommendation can be treated as complete.")

with t3:
    st.subheader("Existing assets")
    assets=remember_table("assets_data",st.data_editor(stored_table("assets_data",DEFAULT_ASSETS),num_rows="dynamic",use_container_width=True,key=table_editor_key("assets_data"),
        column_config={
            "Use for own contribution %":st.column_config.NumberColumn(min_value=0,max_value=100,help="The portion that can be sold or withdrawn to pay the property cost."),
            "Keep invested in wealth plan %":st.column_config.NumberColumn(min_value=0,max_value=100,help="The portion that remains invested and forms the opening investment corpus.")}))
    asset_values=pd.to_numeric(assets["Current value"],errors="coerce").fillna(0)
    contribution_assets=float((asset_values*pd.to_numeric(assets["Use for own contribution %"],errors="coerce").fillna(0)/100).sum())
    usable_assets=float((asset_values*pd.to_numeric(assets["Keep invested in wealth plan %"],errors="coerce").fillna(0)/100).sum())
    overlap=(pd.to_numeric(assets["Use for own contribution %"],errors="coerce").fillna(0)+pd.to_numeric(assets["Keep invested in wealth plan %"],errors="coerce").fillna(0)>100).any()
    if overlap: st.error("An asset cannot be both spent on the property and kept invested. For every row, the two percentages together must not exceed 100%.")
    x,y=st.columns(2); x.metric("Assets available for own contribution",money(contribution_assets)); y.metric("Opening investment corpus",money(usable_assets))
    min_liquidity=st.number_input("Minimum cash liquidity to keep in hand",0.0,value=300000.0,step=25000.0,key="min_liquidity",
        help="Cash reserve that should remain untouched after every property payment and strategy action.")
    bonus_sources_default=pd.DataFrame({"Source":["Annual bonus / incentive"],"Expected date":[(pd.Timestamp(loan_start)+pd.DateOffset(months=5)).date()],"Amount":[0.0]})
    bonus_sources=remember_table("bonus_sources_data",st.data_editor(stored_table("bonus_sources_data",bonus_sources_default),num_rows="dynamic",use_container_width=True,key=table_editor_key("bonus_sources_data"),
        column_config={"Expected date":st.column_config.DateColumn(),"Amount":st.column_config.NumberColumn(min_value=0,format="₹ %.0f",help="A one-time inflow that can be used toward the property contribution.")}))
    st.subheader("Investment assumptions")
    st.caption("Enter the actual holdings/options you want evaluated. Returns are projections, not guarantees.")
    investments=remember_table("investments_data",st.data_editor(stored_table("investments_data",DEFAULT_INVESTMENTS),num_rows="dynamic",use_container_width=True,key=table_editor_key("investments_data")))
    inv_r=weighted_return(investments)
    st.metric("Weighted expected post-tax return",f"{inv_r*100:.2f}%")

property_dates=normalise_schedule(property_schedule.rename(columns={"Expected date":"Date"}),loan_start,"Own contribution")
loan_dates=normalise_schedule(property_schedule.rename(columns={"Expected date":"Date"}),loan_start,"Loan disbursement")
source_dates=normalise_schedule(bonus_sources.rename(columns={"Expected date":"Date"}),loan_start,"Amount")
own_due_by_month=property_dates.groupby("Month")["Own contribution"].sum().to_dict()
loan_by_month=loan_dates.groupby("Month")["Loan disbursement"].sum().to_dict()
source_by_month=source_dates.groupby("Month")["Amount"].sum().to_dict()
last_contribution_month=max(own_due_by_month.keys(),default=1)

def contribution_projection(monthly_saving):
    fund=max(0,contribution_assets-min_liquidity); rows=[]; worst_gap=0.0
    for m in range(1,last_contribution_month+1):
        monthly_available=monthly_saving*(1+growth/100)**((m-1)//12)
        source=source_by_month.get(m,0.0); due=own_due_by_month.get(m,0.0)
        opening=fund; fund+=monthly_available+source-due
        gap=max(0,-fund); worst_gap=max(worst_gap,gap); fund=max(0,fund)
        rows.append({"Month":m,"Calendar month":pd.Timestamp(loan_start)+pd.DateOffset(months=m-1),
                     "Opening contribution fund":opening,"Monthly saving required":monthly_available,
                     "One-time source":source,"Own contribution due":due,"Closing contribution fund":fund,
                     "Protected liquidity":min_liquidity,"Funding gap":gap})
    return worst_gap,rows

lo,hi=0.0,max(own_contribution,safe_surplus*10,1.0)
for _ in range(80):
    mid=(lo+hi)/2
    if contribution_projection(mid)[0]>0: lo=mid
    else: hi=mid
required_monthly_contribution=hi
contribution_gap=max(0,required_monthly_contribution-safe_surplus)
contribution_reserve={m:min(safe_surplus*(1+growth/100)**((m-1)//12),required_monthly_contribution*(1+growth/100)**((m-1)//12)) for m in range(1,last_contribution_month+1)}
_,contribution_rows=contribution_projection(min(required_monthly_contribution,safe_surplus))

payment=safe_surplus+annual_prepay/12
specs=[
    ("Only EMI",0.0,None),("Monthly prepayment",0.0,"monthly"),("Quarterly prepayment",0.0,"quarterly"),
    ("Annual prepayment",0.0,"annual"),("Invest everything",1.0,None),("50% invest + 50% prepay",0.5,"monthly"),
    ("Conservative blend",0.25,"monthly")]
strategy_meanings={
    "Only EMI":"Pay only the scheduled EMI and retain the available surplus as cash.",
    "Monthly prepayment":"Use the available surplus to reduce the loan every month.",
    "Quarterly prepayment":"Accumulate the available surplus and reduce the loan every quarter.",
    "Annual prepayment":"Accumulate the available surplus and reduce the loan once a year.",
    "Invest everything":"Invest all available surplus; do not make voluntary loan prepayments.",
    "50% invest + 50% prepay":"Split the available surplus equally between investing and loan prepayment.",
    "Conservative blend":"Use 75% of available surplus for monthly loan prepayment and invest 25%. This prioritises debt reduction and lower risk while retaining some long-term growth potential."
}
results=[simulate(n,loan,rate,tenure_y*12,payment,usable_assets,inv_r,s,f,growth,leveraged,loan_by_month,contribution_reserve,repayment_mode) for n,s,f in specs]
baseline=results[0]

def score(r):
    if objective=="Close loan fastest": return -r.close_month
    if objective=="Maximise net worth": return r.net_worth
    if objective=="Lowest risk": return -r.risk*1e7+r.net_worth
    if objective=="Highest liquidity": return r.liquidity*1e7+r.net_worth
    return r.net_worth-r.interest*0.4-r.risk*100000

best=max(results,key=score)
with t4:
    st.subheader("Own contribution funding plan")
    f1,f2,f3,f4=st.columns(4)
    f1.metric("Total own contribution",money(own_contribution))
    f2.metric("Usable existing assets",money(max(0,contribution_assets-min_liquidity)))
    f3.metric("Required monthly saving",money(required_monthly_contribution))
    f4.metric("Monthly funding gap",money(contribution_gap))
    if contribution_gap<=1 and abs(entered_own-own_contribution)<=1:
        st.success(f"Achievable on the entered dates. Reserve {money(required_monthly_contribution)} per month for the property first; only the remaining safe surplus should be invested or prepaid.")
    else:
        realistic_extra=math.ceil((contribution_gap*max(1,last_contribution_month))/max(1,safe_surplus))
        st.error(f"The current dates are not fully funded. The plan needs {money(contribution_gap)} more per month. Consider adding eligible savings/FDs, assigning a bonus, reducing the property cost, increasing the permitted loan, cutting expenses, or moving the affected payment date by roughly {realistic_extra} month(s).")
    funding_sources=pd.DataFrame([
        ["Existing assets released for property",max(0,contribution_assets-min_liquidity),"Available now after protecting minimum cash"],
        ["One-time bonus / other sources",sum(source_by_month.values()),"Available only on the entered dates"],
        ["Monthly income savings",required_monthly_contribution,"Reserve this first until the final own-contribution date"],
        ["Minimum cash kept untouched",min_liquidity,"Not counted toward the property payment"]
    ],columns=["Funding source","Amount","Instruction"])
    st.dataframe(funding_sources.style.format({"Amount":money}),use_container_width=True,hide_index=True)
    contribution_view=pd.DataFrame(contribution_rows)
    if not contribution_view.empty:
        st.dataframe(contribution_view.style.format({c:money for c in ["Opening contribution fund","Monthly saving required","One-time source","Own contribution due","Closing contribution fund","Protected liquidity","Funding gap"]}),use_container_width=True,hide_index=True)

    st.markdown(f"<div class='best'><b>Recommended for “{objective}”:</b> {best.name}<br>Allocate {next(s for n,s,f in specs if n==best.name)*100:.0f}% of the calculated surplus to investments and the remainder to prepayment.</div>",unsafe_allow_html=True)
    st.caption(strategy_meanings[best.name])
    with st.expander("What does each strategy mean?"):
        for strategy_name,meaning in strategy_meanings.items():
            st.markdown(f"**{strategy_name}:** {meaning}")
    table=pd.DataFrame([{"Strategy":r.name,"Meaning":strategy_meanings[r.name],"Total interest":r.interest,"Interest saved":baseline.interest-r.interest,"Loan closes in (years)":r.close_month/12,"Investment corpus (30y)":r.corpus,"Projected net worth":r.net_worth,"Liquidity / 10":r.liquidity,"Risk / 10":r.risk} for r in results]).sort_values("Projected net worth",ascending=False)
    st.dataframe(table.style.format({"Total interest":money,"Interest saved":money,"Loan closes in (years)":"{:.1f}","Investment corpus (30y)":money,"Projected net worth":money,"Liquidity / 10":"{:.1f}","Risk / 10":"{:.1f}"}),use_container_width=True)
    chosen=st.selectbox("View action plan",[r.name for r in results],index=[r.name for r in results].index(best.name))
    selected=next(r for r in results if r.name==chosen); schedule=pd.DataFrame(selected.rows)
    monthly_buffer=total_income*buffer_pct/100
    schedule["Protected cash added"]=schedule["Month"].apply(lambda m:monthly_buffer*(1+growth/100)**((m-1)//12))
    schedule["Projected cash liquidity"]=min_liquidity+schedule["Protected cash added"].cumsum()
    schedule["Liquidity surplus above minimum"]=schedule["Projected cash liquidity"]-min_liquidity
    yearly=schedule[schedule.Month%12==0].copy(); yearly["Year"]=yearly.Month//12
    fig=px.line(yearly,x="Year",y=["Loan outstanding","Investment corpus"],title=f"{chosen}: loan and wealth path")
    st.plotly_chart(fig,use_container_width=True)
    st.subheader("Year-end cash liquidity")
    st.caption("This is the protected cash buffer retained outside investment and prepayment. It excludes market-linked investment values.")
    st.dataframe(yearly[["Year","Projected cash liquidity","Liquidity surplus above minimum"]].style.format({"Projected cash liquidity":money,"Liquidity surplus above minimum":money}),use_container_width=True,hide_index=True)
    first=schedule.head(120).copy(); first["Date/action"]=first.Month.apply(lambda m:f"Month {m}")
    st.dataframe(first[["Date/action","Investment","Prepayment","Interest","Loan outstanding"]].style.format({c:money for c in ["Investment","Prepayment","Interest","Loan outstanding"]}),use_container_width=True)
    def build_excel_report():
        output=io.BytesIO()
        inputs=pd.DataFrame([
            ["Selected objective",objective,"Controls which strategy is recommended"],
            ["Recommended strategy",best.name,"Highest-ranked strategy for the selected objective"],
            ["Monthly household income",total_income,"Salary plus spouse/other income"],
            ["Monthly expenses",total_exp,"Total entered recurring expenses"],
            ["Other monthly EMIs",existing_emi,"Debt repayments excluding this home loan"],
            ["Safety buffer %",buffer_pct,"Protected portion of income"],
            ["Calculated safe monthly surplus",safe_surplus,"Income − expenses − other EMIs − safety buffer"],
            ["Property value",property_value,"Entered property price"],
            ["Registration and other charges",charges,"Purchase costs outside property price"],
            ["Home loan amount",loan,"Opening loan principal"],
            ["Home loan interest rate %",rate,"Annual rate used on monthly reducing balance"],
            ["Remaining tenure years",tenure_y,"Remaining contractual loan period"],
            ["Calculated EMI",base_emi,"Monthly reducing-balance EMI"],
            ["Own contribution required",own_contribution,"Property value plus charges minus sanctioned loan"],
            ["Required monthly own-contribution saving",required_monthly_contribution,"Reserved before investment or loan prepayment"],
            ["Minimum cash liquidity",min_liquidity,"Protected and excluded from the property contribution"],
            ["Staged repayment method",repayment_mode,"Controls pre-EMI versus full EMI before final disbursement"],
            ["Expected post-tax investment return %",inv_r*100,"Weighted return after entered tax assumptions"],
            ["Annual income growth %",growth,"Applied to future monthly capacity"],
            ["Loan proceeds invested","Yes" if leveraged else "No","Normally No for a home loan"],
        ],columns=["Input / output","Value","How it affects the plan"])
        with pd.ExcelWriter(output,engine="openpyxl") as writer:
            inputs.to_excel(writer,sheet_name="Inputs & Assumptions",index=False,startrow=2)
            table.to_excel(writer,sheet_name="Strategy Comparison",index=False,startrow=2)
            expenses.to_excel(writer,sheet_name="Expense Details",index=False,startrow=2)
            assets.to_excel(writer,sheet_name="Asset Details",index=False,startrow=2)
            investments.to_excel(writer,sheet_name="Investment Details",index=False,startrow=2)
            property_schedule.to_excel(writer,sheet_name="Property Funding Dates",index=False,startrow=2)
            bonus_sources.to_excel(writer,sheet_name="Contribution Sources",index=False,startrow=2)
            pd.DataFrame(contribution_rows).to_excel(writer,sheet_name="Contribution Plan",index=False,startrow=2)
            for result in results:
                monthly=pd.DataFrame(result.rows)
                monthly.insert(1,"Calendar month",monthly["Month"].apply(lambda m:pd.Timestamp(loan_start)+pd.DateOffset(months=m-1)))
                contribution_monthly=pd.DataFrame(contribution_rows)
                if contribution_monthly.empty:
                    own_paid_by_month={}
                    closing_fund_by_month={}
                else:
                    own_paid_by_month=contribution_monthly.set_index("Month")["Own contribution due"].to_dict()
                    closing_fund_by_month=contribution_monthly.set_index("Month")["Closing contribution fund"].to_dict()
                monthly["Own contribution paid"]=monthly["Month"].map(own_paid_by_month).fillna(0.0)
                monthly["Loan contribution paid"]=monthly["Loan disbursed"]
                monthly["Total property payment"]=monthly["Own contribution paid"]+monthly["Loan contribution paid"]
                monthly["Protected cash added"]=monthly["Month"].apply(lambda m:monthly_buffer*(1+growth/100)**((m-1)//12))
                monthly["Projected cash liquidity"]=min_liquidity+monthly["Protected cash added"].cumsum()
                monthly["Liquidity surplus above minimum"]=monthly["Projected cash liquidity"]-min_liquidity
                monthly["Closing own-contribution fund"]=monthly["Month"].map(closing_fund_by_month).fillna(0.0)
                monthly["Closing cash balance"]=monthly["Projected cash liquidity"]+monthly["Closing own-contribution fund"]
                monthly["Total payment this month"]=monthly["Scheduled EMI"]+monthly["Prepayment"]+monthly["Investment"]
                monthly["Net worth indicator"]=monthly["Investment corpus"]+(loan-monthly["Loan outstanding"])
                sheet_name=result.name[:31]
                monthly.to_excel(writer,sheet_name=sheet_name,index=False,startrow=2)
                ws=writer.book[sheet_name]
                ws["A1"]=f"Monthly action plan — {result.name}"
                ws["A2"]=f"{strategy_meanings[result.name]} Own contribution and loan contribution show property funding; closing cash balance excludes investment value."
            for ws in writer.book.worksheets:
                ws.freeze_panes="A4"
                ws.auto_filter.ref=f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
                ws.sheet_view.showGridLines=False
                if not ws["A1"].value: ws["A1"]=ws.title
                ws["A1"].font=Font(bold=True,size=16,color="174B3A")
                for cell in ws[3]:
                    cell.font=Font(bold=True,color="FFFFFF")
                    cell.fill=PatternFill("solid",fgColor="174B3A")
                for column_cells in ws.columns:
                    letter=column_cells[0].column_letter
                    width=min(42,max(12,max(len(str(c.value or "")) for c in column_cells[:80])+2))
                    ws.column_dimensions[letter].width=width
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value,(int,float)) and any(k in str(ws.cell(3,cell.column).value or "").lower() for k in ["amount","income","expense","emi","interest","principal","balance","prepayment","investment","corpus","worth","value","payment","capacity"]):
                            cell.number_format='₹#,##0.00;[Red]-₹#,##0.00'
                if ws.title not in ["Inputs & Assumptions","Strategy Comparison","Expense Details","Asset Details","Investment Details","Property Funding Dates","Contribution Sources","Contribution Plan"]:
                    for cell in ws[4]:
                        if cell.value=="Calendar month":
                            for c in ws.iter_cols(min_col=cell.column,max_col=cell.column,min_row=4):
                                for x in c: x.number_format="mmm yyyy"
        output.seek(0)
        return output.getvalue()
    st.download_button("Download complete month-wise Excel report",data=build_excel_report(),file_name="wealthpath_month_wise_plan.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    current_values={
        "objective":objective,"leveraged":leveraged,"run_mc":run_mc,
        "salary":salary,"spouse":spouse,"growth":growth,"existing_emi":existing_emi,"buffer_pct":buffer_pct,
        "property_value":property_value,"loan":loan,"charges":charges,"rate":rate,"tenure_y":tenure_y,
        "annual_prepay":annual_prepay,"loan_start":loan_start,"repayment_mode":repayment_mode,
        "expenses_data":expenses,"property_schedule_data":property_schedule,"assets_data":assets,
        "min_liquidity":min_liquidity,"bonus_sources_data":bonus_sources,"investments_data":investments
    }
    current_backup=planner_payload(current_values)
    st.subheader("Save or transfer your entered data")
    st.caption("JSON restores the complete planner exactly. CSV contains the same inputs in an editable import format. Calculated results will be regenerated when imported.")
    e1,e2=st.columns(2)
    e1.download_button("Export complete data as JSON",json.dumps(current_backup,ensure_ascii=False,indent=2).encode("utf-8"),"wealthpath_complete_backup.json","application/json",use_container_width=True)
    e2.download_button("Export entered data as CSV",payload_to_csv(current_backup),"wealthpath_entered_data.csv","text/csv",use_container_width=True)
    if run_mc:
        rng=np.random.default_rng(42); sims=10000
        returns=np.clip(rng.normal(inv_r,0.05,sims),-0.20,0.30)
        corpus=np.array([usable_assets*(1+x/12)**360+payment*((1+x/12)**360-1)/(x/12) if abs(x)>1e-9 else usable_assets+payment*360 for x in returns])
        q=np.percentile(corpus,[10,50,90]); st.subheader("Monte Carlo range")
        a,b,c=st.columns(3); a.metric("Weak case (10th percentile)",money(q[0])); b.metric("Expected case (median)",money(q[1])); c.metric("Strong case (90th percentile)",money(q[2]))

st.divider()
st.caption("Educational planning tool—not personalised financial advice. Review taxes, prepayment rules, floating rates and market assumptions with qualified professionals.")
